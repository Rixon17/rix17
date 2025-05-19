import os
import tempfile
import shutil
import subprocess
from datetime import datetime
import pytz
from github import Github
import pandas as pd
from openpyxl.styles import PatternFill, Border, Side
from concurrent.futures import ThreadPoolExecutor, as_completed
from tqdm import tqdm
import logging
import multiprocessing

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# === CONFIGURATION ===
ORG_NAME = os.getenv('ORG_NAME')
SEARCH_STRING = os.getenv('SEARCH_STRING')
GITHUB_TOKEN = os.getenv('GITHUB_TOKEN')
MAX_WORKERS = max(1, min(multiprocessing.cpu_count() - 1, 15))

# Output file setup
output_dir = os.path.join(os.getcwd(), "Outputs", "Extras")
os.makedirs(output_dir, exist_ok=True)
output_file_path = os.path.join(output_dir, "Data_collected.xlsx")

def git_grep_search(repo_dir, search_string):
    """
    Use git grep for faster searching
    Returns list of (file_path, line_number, line_content)
    """
    results = []
    try:
        # Use git grep with line numbers and showing the matched line
        cmd = [
            "git", "-C", repo_dir,
            "grep", "-n",      # Show line numbers
            "-I",             # Ignore binary files
            "-l",            # Show only filenames
            "-F",            # Fixed strings (don't interpret pattern as regex)
            search_string
        ]
        
        # Get matching files first
        files_output = subprocess.check_output(cmd, text=True).strip()
        
        if files_output:
            for file_path in files_output.split('\n'):
                if file_path:
                    # Now get line numbers and content for each file
                    line_cmd = [
                        "git", "-C", repo_dir,
                        "grep", "-n",
                        "-I",
                        "-F",
                        search_string,
                        file_path
                    ]
                    try:
                        line_output = subprocess.check_output(line_cmd, text=True).strip()
                        for line in line_output.split('\n'):
                            if ':' in line:
                                line_num, content = line.split(':', 1)
                                results.append((file_path, int(line_num), content.strip()))
                    except subprocess.CalledProcessError:
                        continue
                        
    except subprocess.CalledProcessError:
        # git grep returns exit code 1 if no matches found
        pass
    except Exception as e:
        logger.error(f"Error in git grep: {str(e)}")
        
    return results

def clone_and_search_branch(repo_name, branch_name, repo_url):
    """Clone a branch and search for the string"""
    tmpdir = tempfile.mkdtemp()
    results = []
    
    try:
        # Clone specific branch
        subprocess.run([
            "git", "clone",
            "--depth", "1",
            "--single-branch",
            "--branch", branch_name,
            repo_url, tmpdir
        ], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        
        # Search using git grep
        matches = git_grep_search(tmpdir, SEARCH_STRING)
        
        # Format results
        for file_path, line_num, _ in matches:
            link = f"https://github.com/{ORG_NAME}/{repo_name}/blob/{branch_name}/{file_path}#L{line_num}"
            results.append([
                ORG_NAME,
                repo_name,
                branch_name,
                file_path,
                link
            ])
            
    except Exception as e:
        logger.error(f"Error processing {repo_name}/{branch_name}: {str(e)}")
    finally:
        try:
            shutil.rmtree(tmpdir, ignore_errors=True)
        except Exception as e:
            logger.error(f"Failed to delete {tmpdir}: {str(e)}")
            
    return results

def process_repo(repo):
    """Process all branches in a repository"""
    results = []
    repo_url = repo.clone_url.replace("https://", f"https://{GITHUB_TOKEN}@")
    
    try:
        branches = list(repo.get_branches())
        with ThreadPoolExecutor(max_workers=3) as branch_executor:
            branch_futures = [
                branch_executor.submit(clone_and_search_branch, repo.name, branch.name, repo_url)
                for branch in branches
            ]
            
            for future in as_completed(branch_futures):
                results.extend(future.result())
    except Exception as e:
        logger.error(f"Error processing repo {repo.name}: {str(e)}")
        
    return results

def save_to_excel(data, metadata, file_path):
    """Save results to Excel with metadata and styling"""
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        # Metadata sheet
        metadata_df = pd.DataFrame(metadata.items(), columns=['Metadata', 'Value'])
        metadata_df.to_excel(writer, sheet_name='Metadata', index=False)
        
        # Results sheet
        if data:  # Check if we have any results
            results_df = pd.DataFrame(data, columns=["Org", "Repo", "Branch", "File", "Link"])
        else:
            # Create empty DataFrame with columns if no results
            results_df = pd.DataFrame(columns=["Org", "Repo", "Branch", "File", "Link"])
            
        results_df.to_excel(writer, sheet_name="Results", index=False)
        
        # Style both sheets
        for sheet_name in ['Metadata', 'Results']:
            worksheet = writer.sheets[sheet_name]
            
            lighter_blue_fill = PatternFill(start_color='006699', end_color='006699', fill_type='solid')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Style headers
            for cell in worksheet[1]:
                cell.fill = lighter_blue_fill
                cell.border = thin_border
            
            # Style all cells and optimize column widths
            max_lengths = {}
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                                         min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_border
                    col = cell.column
                    max_lengths[col] = max(
                        max_lengths.get(col, 0),
                        len(str(cell.value or ""))
                    )
            
            # Set column widths
            for col, max_length in max_lengths.items():
                worksheet.column_dimensions[col].width = min(max_length + 2, 100)
    
    logger.info(f"Data saved to {file_path}")

def main():
    if not all([GITHUB_TOKEN, ORG_NAME, SEARCH_STRING]):
        logger.error("Missing required environment variables. Please ensure GITHUB_TOKEN, ORG_NAME, and SEARCH_STRING are set.")
        raise ValueError("Missing required environment variables")
    
    # Initialize GitHub client and get user info
    gh = Github(GITHUB_TOKEN)
    try:
        current_user = gh.get_user().login
    except Exception as e:
        logger.error(f"Failed to get GitHub user: {str(e)}")
        current_user = "Unknown"
    
    current_time = datetime.now(pytz.UTC)
    
    metadata = {
        "Current Date and Time (UTC - YYYY-MM-DD HH:MM:SS formatted)": current_time.strftime('%Y-%m-%d %H:%M:%S'),
        "Current User's Login": current_user,
        "Organization": ORG_NAME,
        "Search String": SEARCH_STRING,
        "Worker Count": MAX_WORKERS
    }
    
    logger.info(f"Started search at: {current_time.strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info(f"User: {current_user}")
    logger.info(f"Organization: {ORG_NAME}")
    logger.info(f"Search String: {SEARCH_STRING}")
    
    # Get repositories
    org = gh.get_organization(ORG_NAME)
    repos = list(org.get_repos())
    total_repos = len(repos)
    
    logger.info(f"Found {total_repos} repositories to process")
    
    all_results = []
    
    # Process repositories in parallel
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = [executor.submit(process_repo, repo) for repo in repos]
        
        with tqdm(total=total_repos, desc="Processing repositories") as pbar:
            for future in as_completed(futures):
                results = future.result()
                all_results.extend(results)
                pbar.update(1)
    
    logger.info(f"Search completed. Found {len(all_results)} matches across {total_repos} repositories")
    
    # Save results
    save_to_excel(all_results, metadata, output_file_path)

if __name__ == "__main__":
    main()
